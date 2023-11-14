import openpyxl
from pathlib import Path

p1 = Path.cwd() / 'attachments'
p2 = Path.cwd() / 'result_attachments'

priceUpdates = {'Garlic':3.07, 'Celery':1.19, 'Lemon':1.27}

originWb = openpyxl.load_workbook(p1 / 'produceSales.xlsx')
sheet = originWb.active

for rowNum in range(2, sheet.max_row + 1):
    productName = sheet.cell(rowNum, 1).value
    if productName in priceUpdates.keys():
        sheet.cell(rowNum, 2).value = priceUpdates[productName]

originWb.save(p2 / 'updatedProduceSales.xlsx')